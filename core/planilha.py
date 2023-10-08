import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
import pandas as pd

# Importe os modelos que deseja exportar
from .models import Laboratorio, Equipamento, RegimentoInterno  # Adicione outros modelos conforme necessário

def export_to_excel(request):
    # Consulte os dados que deseja exportar
    laboratorios = Laboratorio.objects.all()
    equipamentos = Equipamento.objects.all()
    regimentos_internos = RegimentoInterno.objects.all()

    # Crie DataFrames a partir dos modelos
    laboratorios_df = pd.DataFrame(list(laboratorios.values()))
    equipamentos_df = pd.DataFrame(list(equipamentos.values()))
    regimentos_internos_df = pd.DataFrame(list(regimentos_internos.values()))

    # Crie um novo arquivo Excel e adicione as planilhas
    workbook = openpyxl.Workbook()
    laboratorios_sheet = workbook.active
    laboratorios_sheet.title = 'Laboratórios'

    # Adicione os dados dos DataFrames às planilhas
    for row in dataframe_to_rows(laboratorios_df, index=False, header=True):
        laboratorios_sheet.append(row)

    equipamentos_sheet = workbook.create_sheet(title='Equipamentos')
    for row in dataframe_to_rows(equipamentos_df, index=False, header=True):
        equipamentos_sheet.append(row)

    regimentos_internos_sheet = workbook.create_sheet(title='Regimentos Internos')
    for row in dataframe_to_rows(regimentos_internos_df, index=False, header=True):
        regimentos_internos_sheet.append(row)

    # Crie uma resposta HTTP para o download do arquivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'

    # Salve o arquivo Excel na resposta HTTP
    workbook.save(response)

    return response
