from django.shortcuts import render, get_object_or_404, redirect, HttpResponseRedirect
from django.urls import reverse

from django.utils.html import strip_tags
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Laboratorio, Infraestrutura,LaboratorioInfraestrutura,ImagemLaboratorio, Marca, Equipamento, RegimentoInterno, UnidadeAcademica,ImagemInfraestrutura,GrupoDePesquisa,MembroLaboratorio
from django.contrib.messages import constants
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.core.exceptions import ObjectDoesNotExist
from PIL import Image
from datetime import datetime
import io
import base64
from django.core.files.base import ContentFile
import logging
from django.templatetags.static import static
from django.core.exceptions import ValidationError
from django.http import FileResponse
from django.core.files import File
from .Send import Email
from .planilha import export_to_excel
from .forms import ExcluiRegimentoInternoForm, ProjetoForm
import cx_Oracle
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
import pandas as pd
from openpyxl.styles import PatternFill
from .models import Unidade, Laboratorio, Equipamento, RegimentoInterno, Unidade 
 # Adicione outros modelos conforme necessário
from django.db.models import Q, Max


from django.db import connections
from .models import Projeto
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import Projeto, Laboratorio
from .forms import ProjetoForm
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt




def visualizar_pdf(request, unidade_academica_id):
    try:
        unidade_academica = UnidadeAcademica.objects.get(id=unidade_academica_id)
        
        if unidade_academica and unidade_academica.pdf:
            pdf_file = unidade_academica.pdf
            response = FileResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{pdf_file.name}"'
            return response
    except UnidadeAcademica.DoesNotExist:
        pass

    return HttpResponse("PDF não encontradoooo.", status=404)

def visualizar_regimento_interno(request, regimento_id):
    try:
        regimento = RegimentoInterno.objects.get(id=regimento_id)

        if regimento and regimento.pdf:
            pdf_file = regimento.pdf
            response = FileResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{pdf_file.name}"'
            return response
    except RegimentoInterno.DoesNotExist:
        pass

    return HttpResponse("Regimento Interno não encontrado.", status=404)

def editar_regimentos_internos(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    regimentos_internos = RegimentoInterno.objects.filter(laboratorio=laboratorio)

    if request.method == 'POST':
        regimento_pdf = request.FILES.get('regimento_pdf')
        
        if regimento_pdf:
            regimento_interno = RegimentoInterno(
                laboratorio=laboratorio,
                pdf=regimento_pdf
            )
            regimento_interno.save()
            return redirect('editar_regimentos_internos', laboratorio_id=laboratorio_id)

    return render(request, 'editar_regimentos_internos.html', {
        'laboratorio': laboratorio,
        'regimentos_internos': regimentos_internos,
    })
def excluir_regimento_interno(request, regimento_id):
    print("View excluir_regimento_interno foi chamada.")
    regimento = get_object_or_404(RegimentoInterno, id=regimento_id)

    # Verifica se o método da solicitação é POST
    if request.method == 'POST':
        # Obtém os dados do formulário
        form = ExcluiRegimentoInternoForm(request.POST)

        # Valida os dados do formulário
        if form.is_valid():
            # Exclui o regimento interno
            regimento.delete()

            # Redireciona de volta para a página de edição de regimentos internos
            return redirect('editar_regimentos_internos', laboratorio_id=regimento.laboratorio.id)

    # Redireciona de volta para a página de edição de regimentos internos
    return redirect('editar_regimentos_internos', laboratorio_id=regimento.laboratorio.id)


def editar_unidades_academicas(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    unidades_academicas = UnidadeAcademica.objects.filter(laboratorio=laboratorio)

    if request.method == 'POST':
        # Lógica para processar o envio da Unidade Acadêmica aqui, se necessário
        unidade_academica_pdf = request.FILES.get('unidade_academica_pdf')
        if unidade_academica_pdf:
            UnidadeAcademica.objects.create(laboratorio=laboratorio, pdf=unidade_academica_pdf)

    return render(request, 'editar_unidades_academicas.html', {
        'laboratorio': laboratorio,
        'unidades_academicas': unidades_academicas,
    })


def visualizar_laboratorio(request, laboratorio_id):
    
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)

    regimentos_internos = RegimentoInterno.objects.filter(laboratorio=laboratorio)

    unidades_academicas = UnidadeAcademica.objects.filter(laboratorio=laboratorio)

    infraestruturas = Infraestrutura.objects.filter(laboratorio_id=laboratorio.id).filter(status=True)

    projetos = Projeto.objects.filter(laboratorio=laboratorio)

    membros = MembroLaboratorio.objects.filter(laboratorio=laboratorio)

    grupos_de_pesquisa = laboratorio.grupos_de_pesquisa.all()

    return render(request, 'view.html', {'laboratorio': laboratorio, 'infraestruturas': infraestruturas, 'regimentos_internos': regimentos_internos, 'unidades_academicas': unidades_academicas, 'projetos': projetos, 'membros': membros, 'grupos_de_pesquisa': grupos_de_pesquisa})


def index(request):
    return render(request, 'index.html')

# @login_required
def edit(request):
   
    laboratorios = Laboratorio.objects.all()
    marcas = Marca.objects.all()
    equipamentos = Equipamento.objects.all()

    return render(request, 'edit.html', {
        'laboratorios': laboratorios,
        'marcas': marcas,
        'equipamentos': equipamentos,
    })

def imagem_rgb(request, imagem_id):
    try:
        imagem = ImagemLaboratorio.objects.get(pk=imagem_id)
        imagem_bin = imagem.imagem

        response = HttpResponse(content_type="image/jpeg")
        image = Image.open(io.BytesIO(imagem_bin))
        
        # Redimensionar a imagem para o tamanho desejado (600x900)
        # new_size = (600, 900)
        # image = image.resize(new_size, Image.ANTIALIAS)
        
        image = image.convert("RGB")
        image.save(response, format="JPEG")

        return response
    except ImagemLaboratorio.DoesNotExist:
        return HttpResponse("Imagem não encontrada.", status=404)

    
    
from django.db.models import Max

def main(request):
    texto = request.GET.get('filter')

    # Verifica se o usuário tem PADACES 708
    padaces_aceitos = []

    if request.user.is_authenticated:
        # Obtém o ID do usuário
        user_id = request.session['user_id'].split(',')[0].split('=')[1]

        # Adicione a informação do USER_LDAP do usuário atual
        user_ldap_session = request.session['user_id'].split(',')[0].split('=')[1]
        print("DEBUG: USER_LDAP do usuário atual--------:", user_ldap_session)
        print("DEBUG: USER_LDAP do usuário atual--------:", user_ldap_session)
        print("DEBUG: USER_LDAP do usuário atual--------:", user_ldap_session)


        # Configurações de conexão com o banco de dados Oracle
        db_settings = {
            'USER': 'cons_oberon',
            'PASSWORD': 'pwdconsoberon',
            'HOST': '10.70.0.14',
            'PORT': '1521',
            'SERVICE_NAME': 'prouea2',
        }

        # Estabelece a conexão com o banco de dados Oracle
        connection = cx_Oracle.connect(
            f"{db_settings['USER']}/{db_settings['PASSWORD']}@{db_settings['HOST']}:{db_settings['PORT']}/{db_settings['SERVICE_NAME']}"
        )

        cursor = connection.cursor()

        try:
            cursor.execute(f"SELECT * FROM OBERON.USUARIOPADACES WHERE USUARIO = '{user_id}' AND PADACES = 708")
            padaces_aceitos = cursor.fetchone()

        finally:
            cursor.close()
            connection.close()

    # Recuperar laboratórios conforme sua lógica de pesquisa
    laboratorios = Laboratorio.objects.all()

    if not padaces_aceitos:
        # Se o usuário não tem PADACES 708, filtre os laboratórios com base no USER_LDAP
        user_ldap = request.session.get('user_ldap', '')

        # Filtre os laboratórios com base no USER_LDAP do responsável e dos responsáveis associados
        user_ldap_session = request.session['user_id'].split(',')[0].split('=')[1]
        laboratorios_filtrados = laboratorios.filter(
            Q(user_ldap_responsavel__iexact=user_ldap_session) |
            Q(responsaveis_associados__user_ldap=user_ldap_session)
        )



        laboratorios = laboratorios_filtrados

    # Adicione estas linhas para depuração
    print('#########################################')
    print("DEBUG: Usuário logado:", request.user.is_authenticated)
    print("DEBUG: ID do usuário logado:", user_id)

    # Verifica se o usuário tem PADACES 708
    is_admin = padaces_aceitos is not None

    for laboratorio in laboratorios:
        # Obtém o USER_LDAP do responsável do laboratório, se existir
        user_ldap_responsavel_laboratorio = ''
        if is_admin and laboratorio.user_ldap_responsavel:
            user_ldap_responsavel_laboratorio = laboratorio.user_ldap_responsavel

            print('-------------------------------------------------------------------')
            print(f"DEBUG: Responsável do laboratório***: {laboratorio.responsavel}")
            print(f"DEBUG: USER_LDAP do responsável do laboratório**: {user_ldap_responsavel_laboratorio}")
            print('-----------------------------------------------------------------------')
            print('-----------------------------------------------------------------------')

    if texto:
        laboratorios = laboratorios.filter(nome_laboratorio__icontains=texto)

    
    is_admin = padaces_aceitos is not None

    if request.method == "POST":
        nome_laboratorio = request.POST.get('nome_laboratorio', '')
        responsavel = request.POST.get('responsavel', '')
        unidade = request.POST.get('unidade', '')
        projeto = request.POST.get('projeto', '')

        # Garanta que os valores não sejam None antes de incluí-los na consulta
        conditions = Q()

        if nome_laboratorio:
            conditions |= Q(nome_laboratorio__icontains=nome_laboratorio)

        if responsavel:
            conditions |= Q(responsavel__icontains=responsavel)

        if unidade:
            conditions |= Q(unidade__icontains=unidade)

        laboratorios = laboratorios.filter(conditions)

        # if projeto:
        #     projetos = Projeto.objects.filter(laboratorio_in=laboratorios, nome_projeto_icontains=projeto)

    else:
        if texto:
            laboratorios = laboratorios.filter(
                Q(nome_laboratorio__icontains=texto) | 
                Q(responsavel__icontains=texto) | 
                Q(bairro__icontains=texto) | 
                Q(grupos_de_pesquisa_nome_do_grupo_icontains=texto) |
                Q(unidade__icontains=texto)
            )
        else:
            laboratorios = Laboratorio.objects.all()
    paginator = Paginator(laboratorios, 6)
    page = request.GET.get('page')
    laboratorios = paginator.get_page(page)

    # Adicione user_ldap_responsavel_laboratorio ao contexto
    context = {
        'laboratorios': laboratorios,
        'is_admin': is_admin,
        'user_ldap_responsavel': user_ldap_session,
        'user_ldap_responsavel_laboratorio': user_ldap_responsavel_laboratorio,
    }

    return render(request, 'main.html', context)
def salvar_laboratorio(request):
    try:
        connection = cx_Oracle.connect(
            user='cons_oberon',
            password='pwdconsoberon',
            dsn='10.70.0.14:1521/prouea2',
            encoding='UTF-8',
            nencoding='UTF-8'
        )

        cursor = connection.cursor()
        user_id = request.session['user_id'].split(',')[0].split('=')[1]

        cursor.execute(f"SELECT * FROM OBERON.USUARIOPADACES WHERE USUARIO = '{user_id}' AND PADACES = 708")
        padaces_aceitos = cursor.fetchone()

        is_admin = padaces_aceitos is not None

        if request.method == "POST":
            imagens_lab = request.FILES.getlist("imagens_lab[]")
            imagens_salvas = []

            for imagem in imagens_lab:
                imagem_laboratorio = ImagemLaboratorio(imagem=imagem)
                imagem_laboratorio.save()
                imagens_salvas.append(imagem_laboratorio)

            if not imagens_lab:
                caminho_imagem_padrao = 'templates/static/images/generica.png'

                with open(caminho_imagem_padrao, 'rb') as img_padrao:
                    imagem_laboratorio = ImagemLaboratorio(imagem=File(img_padrao))
                    imagem_laboratorio.save()
                    imagens_salvas.append(imagem_laboratorio)

            nome_laboratorio = request.POST.get("nome_laboratorio")
            responsavel = request.POST.get("responsavel")
            cpf_responsavel = request.POST.get('cpf_responsavel')
            user_ldap_responsavel = request.POST.get('user_ldap_responsavel')
            email = request.POST.get("email")
            telefone = request.POST.get("telefone")
            unidade = request.POST.get("unidade")
            rua = request.POST.get("rua")
            numero_rua = request.POST.get("numero_rua")
            cep = request.POST.get("cep")
            bairro = request.POST.get("bairro")

            ato_anexo = request.FILES.get("ato_anexo")
            ato_anexo_content = ato_anexo.read() if ato_anexo else None
            
            descricao = request.POST.get("descricao")
            link_pnipe = request.POST.get("link_pnipe")

            andar = request.POST.get("andar")
            sala = request.POST.get("sala")

            andar = andar if andar else None
            sala = sala if sala else None

            unidade = unidade if unidade else None
            bairro = bairro if bairro else None
            rua = rua if rua else None
            numero_rua = numero_rua if numero_rua else None
            cep = cep if cep else None
            
            descricao = descricao if descricao else None
            link_pnipe = link_pnipe if link_pnipe else None

            equipamento_id = request.POST.get('equipamento')
            marca_id = request.POST.get('marca')
            modelo = request.POST.get('modelo')
            finalidade = request.POST.get('finalidade')

            with connection.cursor() as cursor:
                print(f"CPF do responsável (antes de consultar): {cpf_responsavel}")
                cursor.execute(f"SELECT USER_LDAP FROM OBERON.USUARIO WHERE CPF = '{cpf_responsavel}'")
                user_ldap_responsavel = cursor.fetchone()
                print(f"USER_LDAP do responsável (após a consulta): {user_ldap_responsavel}")

                user_ldap_responsavel = user_ldap_responsavel[0] if user_ldap_responsavel else None
                print(f"USER_LDAP do responsável (depois de salvar): {user_ldap_responsavel}")
                print(f"USER_LDAP do responsável (depois de salvar): {user_ldap_responsavel}")

            Email(email, nome_laboratorio, responsavel)

            try:
                equipamento = Equipamento.objects.get(id=equipamento_id)
            except Equipamento.DoesNotExist:
                equipamento = None
            try:
                marca = Marca.objects.get(id=marca_id)
            except Marca.DoesNotExist:
                marca = None

            equipamento = equipamento if equipamento else None
            marca = marca if marca else None
            modelo = modelo if modelo else None
            finalidade = finalidade if finalidade else None

            laboratorio = Laboratorio.objects.create(
                nome_laboratorio=nome_laboratorio,
                responsavel=responsavel,
                email=email,
                telefone=telefone,
                unidade=unidade,
                rua=rua,
                ato_anexo=ato_anexo_content,
                numero_rua=numero_rua,
                cep=cep,
                bairro=bairro,
                
                descricao=descricao,
                link_pnipe=link_pnipe,
                cpf_responsavel=cpf_responsavel,
                user_ldap_responsavel=user_ldap_responsavel,
            )

            pdf_unidade_academica = request.FILES.get("pdf_unidade_academica")
            if pdf_unidade_academica:
                if not pdf_unidade_academica.name.endswith('.pdf'):
                    raise ValidationError("O arquivo deve ser um PDF.")

                unidade_academica = UnidadeAcademica(laboratorio=laboratorio, pdf=pdf_unidade_academica)
                unidade_academica.save()

            infraestrutura = Infraestrutura.objects.create(
                equipamento=equipamento,
                marca=marca,
                modelo=modelo,
                finalidade=finalidade,
            )

            laboratorio.infraestrutura = infraestrutura

            print(f"CPF do responsável (antes de salvar): {cpf_responsavel}")
            print(f"USER_LDAP do responsável (antes de salvar): {user_ldap_responsavel}")

            laboratorio.save()

            print(f"Laboratório salvo. ID: {laboratorio.id}")
            print(f"CPF do responsável (depois de salvar): {laboratorio.cpf_responsavel}")
            print(f"USER_LDAP do responsável (depois de salvar): {laboratorio.user_ldap_responsavel}")

            laboratorio.imagens.set(imagens_salvas)

            if imagens_salvas:
                return HttpResponseRedirect(reverse('editar_laboratorio', args=[laboratorio.id, imagens_salvas[0].id]))
            else:
                caminho_imagem_padrao = static('images/download.jpeg')
                messages.add_message(request, messages.WARNING, 'Nenhuma imagem foi salva.')

                return HttpResponseRedirect(
                    reverse('editar_laboratorio', args=[laboratorio.id, 0]))

    except Exception as e:
        print('Erro:', str(e))
        print('Erro ao salvar')
        messages.add_message(request, messages.WARNING, 'O sistema apresenta falhas internas.')

    finally:
        if connection:
            connection.close()

    return redirect('../edit/')
    

def buscar_nomes(request):
    term = request.GET.get('term', '')

    # Configurações de conexão com o banco de dados Oracle
    db_settings = {
        'USER': 'cons_oberon',
        'PASSWORD': 'pwdconsoberon',
        'HOST': '10.70.0.14',
        'PORT': '1521',
        'SERVICE_NAME': 'prouea2',
    }

    # Estabelece a conexão com o banco de dados Oracle
    connection = cx_Oracle.connect(
        f"{db_settings['USER']}/{db_settings['PASSWORD']}@{db_settings['HOST']}:{db_settings['PORT']}/{db_settings['SERVICE_NAME']}"
    )

    with connection.cursor() as cursor:
        # Consulta SQL para buscar nomes que começam com o termo
        query = f"SELECT NOME_COMPL, CPF FROM OBERON.PESSOA WHERE NOME_COMPL LIKE '{term}%'"
        cursor.execute(query)
        results = cursor.fetchall()

        nomes_com_cpfs = [{'label': row[0], 'value': row[0], 'cpf': row[1]} for row in results]

        # Adiciona USER_LDAP ao resultado
        for pessoa in nomes_com_cpfs:
            cpf = pessoa['cpf']
            # Consulta SQL para buscar USER_LDAP na tabela USUARIO com base no CPF
            cursor.execute(f"SELECT USER_LDAP FROM OBERON.USUARIO WHERE CPF = '{cpf}'")
            user_ldap = cursor.fetchone()
            pessoa['user_ldap'] = user_ldap[0] if user_ldap else None

    connection.close()

    # print("DEBUG: Resultado da busca de nomes:", nomes_com_cpfs)

    return JsonResponse(nomes_com_cpfs, safe=False)





def adicionar_grupo_de_pesquisa(request, laboratorio_id=None):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id) if laboratorio_id else request.user.laboratorio

    if request.method == 'POST':
        nome_do_grupo = request.POST.get('nome_do_grupo')
        area = request.POST.get('area')
        link_grupo = request.POST.get('link_grupo')

        # Crie uma nova instância de GrupoDePesquisa com os dados fornecidos
        grupo_de_pesquisa = GrupoDePesquisa.objects.create(
            nome_do_grupo=nome_do_grupo,
            area=area,
            link_grupo=link_grupo
        )

        # Associe o grupo de pesquisa ao laboratório atual
        laboratorio.grupos_de_pesquisa.add(grupo_de_pesquisa)

        # Redirecione para a página de edição do laboratório ou para onde for adequado
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    grupos_de_pesquisa = laboratorio.grupos_de_pesquisa.all()

    return render(
        request,
        'adicionar_grupo_de_pesquisa.html',
        {'laboratorio': laboratorio, 'grupos_de_pesquisa': grupos_de_pesquisa}
    )

def editar_grupo_de_pesquisa(request, grupo_de_pesquisa_id):
    grupo_de_pesquisa = get_object_or_404(GrupoDePesquisa, id=grupo_de_pesquisa_id)

    if request.method == 'POST':
        grupo_de_pesquisa_id = request.POST.get('grupo_de_pesquisa_id')  # Recupere o ID do grupo de pesquisa
        nome_do_grupo = request.POST.get('nome_do_grupo')
        area = request.POST.get('area')
        link_grupo = request.POST.get('link_grupo')

        # Verifique se o ID no formulário corresponde ao ID do grupo de pesquisa
        if grupo_de_pesquisa_id == str(grupo_de_pesquisa.id):
            # Atualize os valores do grupo de pesquisa
            grupo_de_pesquisa.nome_do_grupo = nome_do_grupo
            grupo_de_pesquisa.area = area
            grupo_de_pesquisa.link_grupo = link_grupo
            grupo_de_pesquisa.save()

            # Redirecione de volta para a página original ou para onde for apropriado
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    # Renderize a página de edição do grupo de pesquisa
    return render(request, 'editar_grupo_de_pesquisa.html', {'grupo_de_pesquisa': grupo_de_pesquisa})


# novo
def excluir_grupo_de_pesquisa(request, grupo_de_pesquisa_id):
    # Obtenha o grupo de pesquisa que você deseja excluir
    grupo_de_pesquisa = get_object_or_404(GrupoDePesquisa, id=grupo_de_pesquisa_id)

    # Lógica para excluir o grupo de pesquisa
    grupo_de_pesquisa.delete()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    
def adicionar_equipamento(request):
   
    if request.method == 'POST':
        nome_equipamento = request.POST.get('nome_equipamento')
        nome_marca = request.POST.get('nome_marca')

        if nome_marca:

            marca_existente = Marca.objects.filter(
                nome_marca=nome_marca
            ).exists()

            if not marca_existente:
                marca = Marca.objects.create(
                    nome_marca=nome_marca
                )
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            else:
                return HttpResponse('Marca já existe.')

        if nome_equipamento:

            equipamento_existente = Equipamento.objects.filter(
                nome_equipamento=nome_equipamento,
            ).exists()

            if not equipamento_existente:
                equipamento = Equipamento.objects.create(
                    nome_equipamento=nome_equipamento,
                )
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            else:
                # Adiciona a mensagem de erro ao sistema de mensagens
                messages.error(request, 'Equipamento já existe.')
   
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def editar_infraestrutura(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
   
    marcas = Marca.objects.all() 
    equipamentos = Equipamento.objects.all()

 
    if request.method == 'POST':
        equipamento_id = request.POST.get('equipamento')
        marca_id = request.POST.get('marca')
        modelo = request.POST.get('modelo')
        finalidade = request.POST.get('finalidade')
        tombo = request.POST.get('tombo')
        quantidade = request.POST.get('quantidade')
        # Verifique se uma imagem foi enviada
        novas_imagens = request.FILES.getlist('nova_imagem')

        # Crie uma nova infraestrutura
        infraestrutura = Infraestrutura.objects.create(
            equipamento_id=equipamento_id,
            marca_id=marca_id,
            laboratorio_id=laboratorio_id,
            modelo=modelo,
            finalidade=finalidade,
            tombo=tombo,
            quantidade=quantidade,
        )

        # Associe a nova infraestrutura ao laboratório atual
        # LaboratorioInfraestrutura.objects.create(laboratorio=laboratorio, infraestrutura=infraestrutura)

        # Se uma nova imagem foi enviada, crie uma instância de ImagemInfraestrutura para cada imagem
        for nova_imagem in novas_imagens:
            imagem_infraestrutura = ImagemInfraestrutura(imagem=nova_imagem, infraestrutura=infraestrutura)
            imagem_infraestrutura.save()

        print('Infraestrutura criada com sucesso!')
        # Redirecione para a página de edição do laboratório ou para onde for adequado
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    
    infraestruturas = Infraestrutura.objects.filter(laboratorio_id=laboratorio_id)
    return render(
        request,
        'editar_infraestrutura.html',
        {'laboratorio': laboratorio, 'marcas': marcas, 'equipamentos': equipamentos, 'infraestruturas': infraestruturas}
    )

def equipamento(request):
   
    laboratorios = Laboratorio.objects.all()
  
    return render(request, 'equipamento.html', {
        'laboratorios': laboratorios,
    })

def salvar_equipamento(request):
    if request.method == "POST":
        try:   
            equipamento = request.POST.get("equipamento")
            marca = request.POST.get("marca")
            modelo = request.POST.get("modelo")
            finalidade = request.POST.get("finalidade")
          
            
  # Salvar os dados no banco de dados
            Infraestrutura.objects.create(
                equipamento=equipamento,
                marca=marca,
                modelo=modelo,
                finalidade=finalidade,
                
            )

            messages.add_message(request, messages.SUCCESS, 'Salvo com sucesso.')
            return render(request, 'equipamento.html')
        except:
            messages.add_message(request, messages.WARNING, 'O sistema apresenta falhas internas.')
            return redirect('../equipamento/')
    else:
        return render(request, 'equipamento.html')



def paginar(list, limit_per_page, request): 
    paginator = Paginator(list, limit_per_page) 
    page = request.GET.get('page')
    atos = paginator.get_page(page)
    context = {'atos': atos}
    return context

@login_required
def aprovados(request):
    atos_list = AtoNormativ.objects.filter(status='aprovado')
    context = paginar(atos_list, 16, request);
    return render(request, 'main.html', context)


def pesquisar(request):
    if request.method == 'GET':
        termo_pesquisa = request.GET.get('pesquisa', '')  # Obtém o termo de pesquisa do parâmetro GET 'pesquisa'
        atos = AtoNormativ.objects.filter(texto_normativo__icontains=termo_pesquisa)  # Filtra os atos com base no termo de pesquisa
        context = {'atos': atos}
        return render(request, 'resultado_pesquisa.html', context)
    
def visualizar_arquivo(request, laboratorio_id):
    try:
        laboratorio = Laboratorio.objects.get(pk=laboratorio_id)
        if laboratorio.ato_anexo:
            response = HttpResponse(laboratorio.ato_anexo, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="arquivo.pdf"'  # Pode ajustar o nome do arquivo
            return response
    except Laboratorio.DoesNotExist:
        pass

    return HttpResponse("Arquivo não encontrado.", status=404)

def visualizar_imagens(request, laboratorio_id):
    laboratorio = Laboratorio.objects.get(pk=laboratorio_id)
    imagens = laboratorio.imagens.all()
    return render(request, 'visualizar_imagens.html', {'laboratorio': laboratorio, 'imagens': imagens})

def view_imagem(request, imagem_id):
    try:
        imagem = ImagemLaboratorio.objects.get(pk=imagem_id)
        imagem_bin = imagem.imagem

        # Configurar o cabeçalho de tipo de conteúdo para uma imagem
        response = HttpResponse(content_type="image/jpeg")
        
        # Abra a imagem usando o Pillow (PIL)
        image = Image.open(io.BytesIO(imagem_bin))

        # Converta a imagem para o modo RGB
        image = image.convert("RGB")
        
        # Salve a imagem no formato JPEG
        image.save(response, format="JPEG")

        return response
    except ObjectDoesNotExist:
        return HttpResponse("Imagem não encontrada.", status=404)

from .models import ResponsavelAssociado

def editar_laboratorio(request, laboratorio_id, imagem_id):

    print("DEBUG: USER_LDAP do responsável do laboratório na editar_laboratorio:", request.session.get('user_ldap', ''))
    print("DEBUG: USER_LDAP do responsável do laboratório na editar_laboratorio:", request.session.get('user_id', ''))

    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)

    # Verifica se o usuário tem PADACES 708
    padaces_aceitos = []

    if request.user.is_authenticated:

        # Obtém o ID do usuário
        user_id = request.session['user_id'].split(',')[0].split('=')[1]

        # Configurações de conexão com o banco de dados Oracle
        db_settings = {
            'USER': 'cons_oberon',
            'PASSWORD': 'pwdconsoberon',
            'HOST': '10.70.0.14',
            'PORT': '1521',
            'SERVICE_NAME': 'prouea2',
        }

        # Estabelece a conexão com o banco de dados Oracle
        connection = cx_Oracle.connect(
            f"{db_settings['USER']}/{db_settings['PASSWORD']}@{db_settings['HOST']}:{db_settings['PORT']}/{db_settings['SERVICE_NAME']}"
        )

        cursor = connection.cursor()

        try:
            # Use o ID do usuário na consulta SQL
            cursor.execute(f"SELECT * FROM OBERON.USUARIOPADACES WHERE USUARIO = '{user_id}' AND PADACES = 708")
            padaces_aceitos = cursor.fetchone()

        finally:
            cursor.close()
            connection.close()

    # Lógica para verificar se o usuário é admin
    is_admin = padaces_aceitos is not None


    # Mensagens de depuração
    print("DEBUG: Usuário logado:", request.user.is_authenticated)
    print("DEBUG: ID do usuário logado:", user_id)
    print("DEBUG: É admin:", is_admin)

    if request.method == 'POST':
        # Processar os campos do formulário
        laboratorio.nome_laboratorio = request.POST.get('nome_laboratorio', '')
        laboratorio.responsavel = request.POST.get('responsavel', '')
        laboratorio.email = request.POST.get('email', '')
        laboratorio.data_criacao = request.POST.get('data_criacao', '')
        laboratorio.descricao = request.POST.get('descricao', '')
        laboratorio.link_pnipe = request.POST.get('link_pnipe', '')
        laboratorio.user_ldap_responsavel = request.POST.get('user_ldap_responsavel', '')

        # Processar os novos responsáveis associados
        novos_responsaveis_associados = request.POST.getlist('novo_responsavel_1', '')

        # Adicione estas variáveis para armazenar os dados do novo responsável associado
        novo_responsavel_cpf = None
        novo_responsavel_user_ldap = None

        for nome_responsavel in novos_responsaveis_associados:
            cpf_responsavel = request.POST.get(f'cpf_responsavel_1', '')
            user_ldap_responsavel = request.POST.get(f'user_ldap_responsavel_1', '')

            responsavel_associado = ResponsavelAssociado.objects.create(
                nome=nome_responsavel,
                cpf=cpf_responsavel,
                user_ldap=user_ldap_responsavel,
                laboratorio=laboratorio
            )
            laboratorio.responsaveis_associados.add(responsavel_associado)

            # Atualize as variáveis com os dados do último responsável associado
            novo_responsavel_cpf = responsavel_associado.cpf
            novo_responsavel_user_ldap = responsavel_associado.user_ldap

            # Adicione estes prints para depuração, agora dentro do loop
            print(f"Novo responsável associado salvo. ID: {responsavel_associado.id}")
            print(f"CPF do novo responsável associado: {novo_responsavel_cpf}")
            print(f"USER_LDAP do novo responsável associado: {novo_responsavel_user_ldap}")

        # Processar a imagem, se presente
        imagem_laboratorio = request.FILES.get('imagem_laboratorio')
        if imagem_laboratorio:
            # Crie uma instância de ImagemLaboratorio associada ao laboratório
            imagem = ImagemLaboratorio(imagem=imagem_laboratorio)
            imagem.save()
            laboratorio.imagens.add(imagem)

        laboratorio.save()

        messages.success(request, 'Laboratório atualizado com sucesso.')
        return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=imagem_id)

        # Restante do código para renderizar a página de edição
    return render(request, 'editar_laboratorio.html', {'laboratorio': laboratorio, 'is_admin': is_admin})

from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from .models import Laboratorio, ResponsavelAssociado
from django.urls import reverse

def excluir_responsavel_associado(request, laboratorio_id, responsavel_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    responsavel_associado = get_object_or_404(ResponsavelAssociado, id=responsavel_id)

    # Certifique-se de que o responsável associado está vinculado ao laboratório antes de excluir
    if responsavel_associado in laboratorio.responsaveis_associados.all():
        responsavel_associado.delete()
        messages.success(request, 'Responsável associado excluído com sucesso.')
    else:
        messages.error(request, 'Erro ao excluir responsável associado.')
        print(f"DEBUG: Responsável associado não encontrado no laboratório {laboratorio.id}")

    # Redirecione de volta para a mesma página
    return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=0)










def editar_endereco(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, pk=laboratorio_id)

    oracle_db_config = {
        'user': 'iury',
        'password': 'iury2023!',
        'dsn': 'sgbd01.uea.br:1521/prouea',
    }

    connection = cx_Oracle.connect(**oracle_db_config)
    cursor = connection.cursor()

    u = f"SELECT * FROM XPROJ2.UNIDADE"
    cursor.execute(u)
    unidade = cursor.fetchall() 

    if request.method == 'POST':
        try:
            # Recupere os valores atualizados dos campos de endereço do formulário
            unidade = request.POST.get('unidade')
            rua = request.POST.get('rua')
            numero_rua = request.POST.get('numero_rua')
            cep = request.POST.get('cep')
            bairro = request.POST.get('bairro')
            andar = request.POST.get('andar')
            sala = request.POST.get('sala')

            # Atualize os campos de endereço do objeto Laboratorio
            laboratorio.unidade = unidade
            laboratorio.rua = rua
            laboratorio.numero_rua = numero_rua
            laboratorio.cep = cep
            laboratorio.bairro = bairro
            laboratorio.andar=andar
            laboratorio.sala=sala

            # Salve o objeto Laboratorio atualizado no banco de dados
            laboratorio.save()

            # Adicione uma mensagem de sucesso
            messages.success(request, 'Endereço atualizado com sucesso.')

            # Redirecione para a página de detalhes do laboratório ou para onde desejar
            # Neste exemplo, estamos redirecionando para a página de edição de endereço novamente
            return redirect('editar_endereco', laboratorio_id=laboratorio_id)
        except Exception as e:
            # Trate qualquer exceção que possa ocorrer
            print('Erro:', str(e))
            messages.error(request, f'Erro ao atualizar o endereço: {str(e)}')
    # Renderize o template com os detalhes do laboratório e o formulário para edição de endereço
    return render(request, 'editar_endereco.html', {'laboratorio': laboratorio, 'unidade': unidade})


def visualizar_membros_laboratorio(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    membros = laboratorio.membros.all()

    if request.method == 'POST':
        # Lógica para adicionar ou atualizar membros, se necessário
        nome_membro = request.POST.get('nome_membro')
        funcao = request.POST.get('funcao')
        curriculo_lattes = request.POST.get('curriculo_lattes')

        if nome_membro and funcao:
            membro, created = MembroLaboratorio.objects.get_or_create(
                laboratorio=laboratorio,
                nome_membro=nome_membro,
                defaults={'funcao': funcao, 'curriculo_lattes': curriculo_lattes}
            )

    return render(
        request,
        'visualizar_membros.html',
        {'laboratorio': laboratorio, 'membros': membros}
    )
def editar_membro_laboratorio(request, membro_id):
    membro = get_object_or_404(MembroLaboratorio, id=membro_id)

    if request.method == 'POST':
        # Obter os dados do formulário
        nome_membro = request.POST.get('nome_membro')
        funcao = request.POST.get('funcao')
        curriculo_lattes = request.POST.get('curriculo_lattes')

        # Atualizar as informações do membro
        membro.nome_membro = nome_membro
        membro.funcao = funcao
        membro.curriculo_lattes = curriculo_lattes
        membro.save()

        # Redirecionar de volta para a página original
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    return render(
        request,
        'editar_membro_laboratorio.html',
        {'membro': membro}
    )


# novo
def excluir_membro_laboratorio(request, membro_id):
    membro = get_object_or_404(MembroLaboratorio, id=membro_id)
    membro.delete()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def export_to_excel(request):

    if request.method == "POST":
        unidade = request.POST.getlist('unidade')
        export_all = request.POST.get('export_all')
        laboratorio = request.POST.getlist('laboratorios')

       
        if unidade:
            if len(unidade) == 1:
                laboratorios = Laboratorio.objects.filter(unidade=unidade[0])
            else:
                laboratorios = Laboratorio.objects.filter(unidade__in=unidade)
          
            equipamentos = Infraestrutura.objects.all()
            regimentos_internos = RegimentoInterno.objects.all()

        if export_all:
            laboratorios = Laboratorio.objects.all()
            equipamentos = Infraestrutura.objects.all()
            regimentos_internos = RegimentoInterno.objects.all()

        if laboratorio:
            laboratorios = Laboratorio.objects.filter(id__in=laboratorio)
            equipamentos = Infraestrutura.objects.filter(laboratorio__in=laboratorios)
            regimentos_internos = RegimentoInterno.objects.all()

    laboratorios_df = pd.DataFrame(list(laboratorios.values()))
    if regimentos_internos.exists():
        regimentos_internos_df = pd.DataFrame(list(regimentos_internos.values()))
    else:
        regimentos_internos_df = pd.DataFrame()
        
    equipamentos_df = pd.DataFrame(list(equipamentos.values()))
    
    # if 'laboratorio_id' in equipamentos_df.columns:
    #     equipamentos_df = equipamentos_df.drop(columns=['laboratorio_id'])

    equipamentos_df['status'] = equipamentos_df['status'].apply(lambda x: 'Ativo' if x == 1 else 'Em Manutenção')
    regimentos_internos_df['status_x'] = regimentos_internos_df['status'].apply(lambda x: 'Ativo' if x == 1 else 'Inativo')
    equipamentos_df['marca_id'] = equipamentos_df['marca_id'].apply(lambda x: Marca.objects.get(id=x).nome_marca if Marca.objects.filter(id=x).exists() else 'Desconhecida') 
    equipamentos_df['equipamento_id'] = equipamentos_df['equipamento_id'].apply(lambda x: Equipamento.objects.get(id=x).nome_equipamento if Equipamento.objects.filter(id=x).exists() else 'Desconhecida') 
    
    # combined_df = pd.merge(laboratorios_df, regimentos_internos_df, left_on='id', right_on='laboratorio_id', how='left')
    # combined_df = pd.merge(combined_df, equipamentos_df, left_on='laboratorio_id', right_on='id', how='left')
    combined_df = pd.merge(laboratorios_df, regimentos_internos_df, left_on='id', right_on='laboratorio_id', how='left')

    equipamentos_grouped = equipamentos_df.groupby('laboratorio_id').agg(list).reset_index()
    combined_df = pd.merge(combined_df, equipamentos_grouped, left_on='laboratorio_id', right_on='laboratorio_id', how='left')

    combined_df['equipamento_id'] = combined_df['equipamento_id'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['marca_id'] = combined_df['marca_id'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['modelo'] = combined_df['modelo'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['finalidade'] = combined_df['finalidade'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['status_y'] = combined_df['status_y'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['tombo'] = combined_df['tombo'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['quantidade'] = combined_df['quantidade'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)

    column_name_mapping = {
        'nome_laboratorio': 'LABORATÓRIO',
        'responsavel': 'RESPONSAVEL',
        'email': 'EMAIL',
        'telefone': 'TELEFONE',
        'unidade': 'UNIDADE',
        'rua': 'RUA',
        'numero_rua': 'NÚMERO',
        'cep': 'CEP',
        'bairro': 'BAIRRO',
        'andar': 'ANDAR',
        'sala': 'SALA',
        
        'descricao': 'DESCRICAO',
        'link_pnipe': 'LINK',
        'ato_anexo': 'ANEXO',
        'pdf': ' PDF',
        'nome_do_pdf': 'NOME_PDF',
        'status_y': 'STATUS_INFRA',
        'status_x': 'STATUS_PDF',

    }

    combined_df = combined_df.rename(columns=column_name_mapping)

    combined_df = combined_df.sort_values(by='id')

    combined_df = combined_df.drop(columns=['id', 'laboratorio_id', 'id_x', 'id_y'])

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.title = 'Laboratório'

    fill_green_dark = PatternFill(start_color="008000", end_color="008000", fill_type="solid")

    row_colors = ['00ff00', '4aea37', '70bf5d', '79aa6b', '7e9576', '808080']

    for col_idx, column_name in enumerate(combined_df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=column_name)
        cell.fill = fill_green_dark

    for row_idx, row in enumerate(combined_df.itertuples(), 2):
        row_color = row_colors[row_idx % len(row_colors)]
        for col_idx, value in enumerate(row[1:], 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="laboratorios_data.xlsx"'

    workbook.save(response)

    return response
    


def delete_infra(request, laboratorio_id):

    Infraestrutura.objects.filter(id=laboratorio_id).delete()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


def change_status(request, id):
    try:
        equipamento = Infraestrutura.objects.get(id=id)
    except Infraestrutura.DoesNotExist:
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    equipamento.status = not equipamento.status
    equipamento.save()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def change_status_pdf(request, id):
    try:
        regimento = RegimentoInterno.objects.get(id=id)
    except Infraestrutura.DoesNotExist:
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    regimento.status = not regimento.status
    regimento.save()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


def projetos(request, laboratorio_id):
    oracle_db_config = {
        'user': 'iury',
        'password': 'iury2023!',
        'dsn': 'sgbd01.uea.br:1521/prouea',
    }

    connection = cx_Oracle.connect(**oracle_db_config)
    cursor = connection.cursor()
    cursor.execute('SELECT PROJETO, TITULO FROM XPROJ2.PROJETO')
    resultados = cursor.fetchall()
    cursor.close()
    connection.close()

    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    projetos = Projeto.objects.filter(laboratorio=laboratorio)

    if request.method == 'POST':
        mutable_post = request.POST.copy()
        mutable_post['laboratorio'] = laboratorio_id
        form = ProjetoForm(mutable_post)

        if form.is_valid():
            projeto = form.save(commit=False)
            projeto.laboratorio = laboratorio
            projeto.save()
            return redirect('projetos', laboratorio_id=laboratorio_id)
    else:
        form = ProjetoForm(initial={'laboratorio': laboratorio_id, 'tem_cadastro_uea': True})

    projetos = Projeto.objects.filter(laboratorio=laboratorio)

    return render(request, 'projetos.html', {'form': form, 'projetos': projetos, 'laboratorio': laboratorio, 'resultados': resultados})


def excluir_projeto(request, projeto_id):
    projeto = get_object_or_404(Projeto, id=projeto_id)
    laboratorio_id = projeto.laboratorio.id  # Obtém o laboratório associado ao projeto
    projeto.delete()
    return redirect('projetos', laboratorio_id=laboratorio_id)  # Redirecione de volta para a página de projetos após excluir


def obter_nome_projeto(request):
    projeto_selecionado = request.GET.get('projeto', '')
    oracle_db_config = {
        'user': 'iury',
        'password': 'iury2023!',
        'dsn': 'sgbd01.uea.br:1521/prouea',
    }

    connection = cx_Oracle.connect(**oracle_db_config)
    cursor = connection.cursor()
    query = f"SELECT TITULO FROM XPROJ2.PROJETO WHERE PROJETO = '{projeto_selecionado}'"
    cursor.execute(query)
    resultado = cursor.fetchone()
    cursor.close()
    connection.close()

    nome_projeto = resultado[0] if resultado else ''
    return JsonResponse({'nome_projeto': nome_projeto})

@csrf_exempt
def obter_detalhes_projeto(request):
    projeto_selecionado = request.GET.get('projeto', '')
    oracle_db_config = {
        'user': 'iury',
        'password': 'iury2023!',
        'dsn': 'sgbd01.uea.br:1521/prouea',
    }

    connection = cx_Oracle.connect(**oracle_db_config)
    cursor = connection.cursor()

    # Consulta para obter o nome do projeto
    query_projeto = f"SELECT TITULO, TIPO, AUTOR FROM XPROJ2.PROJETO WHERE PROJETO = '{projeto_selecionado}'"
    cursor.execute(query_projeto)
    resultado_projeto = cursor.fetchone()

    # Consulta para obter o nome do docente associado ao projeto (autor)
    nome_autor_projeto = resultado_projeto[2] if resultado_projeto and resultado_projeto[2] else ''

    # Consulta para obter os discentes participantes associados ao projeto
    query_discentes = f"SELECT MEMBRO FROM XPROJ2.PROJ_MEMBRO WHERE PROJETO = '{projeto_selecionado}' AND PERFIL IN ('BOL', 'VOL', 'MEM')"
    cursor.execute(query_discentes)
    resultados_matricula_discente = cursor.fetchall()

    # Adiciona a lógica para verificar CPF na tabela EXTERNO
    nomes_discentes = []
    for resultado in resultados_matricula_discente:
        matricula = str(resultado[0])
        if len(matricula) == 14:
            # Se a matrícula tem 14 dígitos, consulte a tabela EXTERNO
            query_cpf_externo = f"SELECT NOME FROM XPROJ2.EXTERNO WHERE CPF = '{matricula}'"
            cursor.execute(query_cpf_externo)
            resultado_cpf_externo = cursor.fetchone()
            if resultado_cpf_externo:
                nomes_discentes.append(resultado_cpf_externo[0])
            else:
                nomes_discentes.append(matricula)
        else:
            nomes_discentes.append(matricula)

    # Consulta para obter a modalidade
    query_tipo_projeto = f"SELECT TIPO FROM XPROJ2.PROJETO WHERE PROJETO = '{projeto_selecionado}'"
    cursor.execute(query_tipo_projeto)
    resultado_tipo_projeto = cursor.fetchone()

    # Se o tipo do projeto foi encontrado, continue para obter a modalidade
    if resultado_tipo_projeto and resultado_tipo_projeto[0]:
        tipo_projeto = resultado_tipo_projeto[0]

        # Consulta para obter a modalidade usando o tipo
        query_modalidade = f"SELECT NOME FROM XPROJ2.TIPO WHERE TIPO = '{tipo_projeto}'"
        cursor.execute(query_modalidade)
        resultado_modalidade = cursor.fetchone()

        modalidade_projeto = resultado_modalidade[0] if resultado_modalidade else ''
    else:
        modalidade_projeto = ''

    # Consulta para obter as matrículas dos discentes participantes
    query_matricula_discente = f"SELECT MEMBRO FROM XPROJ2.PROJ_MEMBRO WHERE PROJETO = '{projeto_selecionado}' AND PERFIL IN ('BOL', 'VOL', 'MEM')"
    cursor.execute(query_matricula_discente)
    resultados_matricula_discente = cursor.fetchall()

    # Modificação para extrair o primeiro elemento de cada tupla
    matricula_discente = ', '.join(
        str(resultado[0]) for resultado in resultados_matricula_discente) if resultados_matricula_discente else ''

    # Consulta para obter o fomento
    query_fomento = f"SELECT CATEGORIA FROM XPROJ2.PROJETO WHERE PROJETO = '{projeto_selecionado}'"
    cursor.execute(query_fomento)
    resultado_fomento = cursor.fetchone()

    # Verificação adicional para lidar com None e converter números para string
    fomento_projeto = str(resultado_fomento[0]) if (
            resultado_fomento and resultado_fomento[0] is not None) else 'Fomento não encontrado no banco de dados.'

    # Adicione esta linha para verificar o valor atribuído
    print('Valor de fomento_projeto:', fomento_projeto)

    cursor.close()
    connection.close()

    nome_projeto = resultado_projeto[0] if resultado_projeto else ''

    # Retorna a resposta JSON com as novas linhas adicionadas
    return JsonResponse({
        'nome_projeto': nome_projeto,
        'nome_autor_projeto': nome_autor_projeto,
        'nomes_discentes': nomes_discentes,
        'modalidade_projeto': modalidade_projeto,
        'matricula_discente': matricula_discente,
        'fomento_projeto': fomento_projeto
    })



def obter_vigencia_projeto(request):
    projeto_selecionado = request.GET.get('projeto', '')
    oracle_db_config = {
        'user': 'iury',
        'password': 'iury2023!',
        'dsn': 'sgbd01.uea.br:1521/prouea',
    }

    connection = cx_Oracle.connect(**oracle_db_config)
    cursor = connection.cursor()

    # Consulta para obter a data de início e término do projeto
    query_vigencia = f"SELECT VIG_DT_INICIO, VIG_DT_TERMINO FROM XPROJ2.PROJETO WHERE PROJETO = '{projeto_selecionado}'"
    cursor.execute(query_vigencia)
    resultado_vigencia = cursor.fetchone()

    cursor.close()
    connection.close()

    vigencia_inicio = resultado_vigencia[0].strftime('%Y-%m-%d') if resultado_vigencia and resultado_vigencia[0] else ''
    vigencia_fim = resultado_vigencia[1].strftime('%Y-%m-%d') if resultado_vigencia and resultado_vigencia[1] else ''

    return JsonResponse({'vigencia_inicio': vigencia_inicio, 'vigencia_fim': vigencia_fim})



def obter_modalidade_projeto(request):
    projeto_selecionado = request.GET.get('projeto', '')
    oracle_db_config = {
        'user': 'iury',
        'password': 'iury2023!',
        'dsn': 'sgbd01.uea.br:1521/prouea',
    }

    connection = cx_Oracle.connect(**oracle_db_config)
    cursor = connection.cursor()

    # Consulta para obter o tipo do projeto
    query_tipo_projeto = f"SELECT TIPO FROM XPROJ2.PROJETO WHERE PROJETO = '{projeto_selecionado}'"
    cursor.execute(query_tipo_projeto)
    resultado_tipo_projeto = cursor.fetchone()

    print(f"Resultado Tipo Projeto: {resultado_tipo_projeto}")  # Adiciona este log

    # Se o tipo do projeto foi encontrado, continue para obter a modalidade
    if resultado_tipo_projeto and resultado_tipo_projeto[0]:
        tipo_projeto = resultado_tipo_projeto[0]

        # Consulta para obter a modalidade usando o tipo
        query_modalidade = f"SELECT NOME FROM XPROJ2.TIPO WHERE TIPO = '{tipo_projeto}'"
        cursor.execute(query_modalidade)
        resultado_modalidade = cursor.fetchone()

        print(f"Resultado Modalidade: {resultado_modalidade}")  # Adiciona este log

        modalidade = resultado_modalidade[0] if resultado_modalidade else ''
    else:
        modalidade = ''

    cursor.close()
    connection.close()

    return JsonResponse({'modalidade': modalidade})
